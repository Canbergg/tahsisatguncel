import streamlit as st
import pandas as pd
from openpyxl import load_workbook

def process_excel(file):
    df = pd.read_excel(file)
    
    # Sütun isimlerini al
    columns = df.columns.tolist()
    
    # AE ve AF sütunlarının gerçek başlıklarını belirle
    ae_column = columns[30]  # AE sütunu (31. sütun, 0 indeksli olduğu için 30)
    af_column = columns[31]  # AF sütunu (32. sütun, 0 indeksli olduğu için 31)
    
    # Yeni sütunları ekleyelim
    df.insert(df.columns.get_loc("MaxNeedForSalesParam"), "İlişki", None)
    df.insert(df.columns.get_loc("MaxNeedForSalesParam"), "Unique Count", None)
    
    # Sütun sırasını düzelterek tekrar oluştur
    column_order = df.columns.tolist()
    column_order.remove("Unique Count")
    column_order.remove("İlişki")
    insert_loc = column_order.index("MaxNeedForSalesParam")
    column_order.insert(insert_loc, "İlişki")
    column_order.insert(insert_loc, "Unique Count")
    df = df[column_order]
    
    # Unique Count hesaplama
    if 'Mağaza Adı' in df.columns:
        unique_store_count = df['Mağaza Adı'].nunique()
        df['Unique Count'] = df[ae_column].map(df[ae_column].value_counts()) / unique_store_count
    else:
        st.error("'Mağaza Adı' sütunu eksik, lütfen dosyanızı kontrol edin.")
    
    # İlişki sütunu doldurma
    if af_column in df.columns:
        df['İlişki'] = df[af_column].apply(lambda x: "Muadil" if x == 11 else ("Muadil Stoksuz" if x == 10 else "İlişki yok"))
    else:
        st.error("AF sütunu eksik, lütfen dosyanızı kontrol edin.")
    
    # Özel sıralama işlemi
    sort_columns = ['Mağaza Adı', 'ItAtt48', 'Ürün Brüt Ağırlık']
    if all(col in df.columns for col in sort_columns):
        df = df.sort_values(by=['Mağaza Adı', 'ItAtt48', 'Ürün Brüt Ağırlık'], ascending=[True, True, True])
    else:
        st.error("Sıralama için gerekli sütunlar eksik: 'Mağaza Adı', 'ItAtt48' veya 'Ürün Brüt Ağırlık'")
    
    return df

st.title("Excel Veri İşleme Uygulaması")
uploaded_file = st.file_uploader("Excel dosyanızı yükleyin", type=["xlsx", "xls"])

if uploaded_file is not None:
    try:
        processed_df = process_excel(uploaded_file)
        
        # Unique Count değerlerine göre ayırma
        tekli_df = processed_df[processed_df['Unique Count'] == 1].copy()
        ciftli_df = processed_df[processed_df['Unique Count'] == 2].copy()
        uclu_df = processed_df[processed_df['Unique Count'] == 3].copy()
        
        # Excel dosyasını birden fazla sayfa olarak kaydetme
        output_file = "processed_data.xlsx"
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            processed_df.to_excel(writer, sheet_name='Tüm Veriler', index=False)
            tekli_df.to_excel(writer, sheet_name='Tekli', index=False)
            ciftli_df.to_excel(writer, sheet_name='Çift', index=False)
            uclu_df.to_excel(writer, sheet_name='Üçlü', index=False)
        
        # Excel dosyasına formülleri ekleme
        wb = load_workbook(output_file)
        ws = wb['Çift']
        ws['AQ1'] = "İhtiyaç"
        ws['AR1'] = "İhtiyaç Çoklama"
        ws['AS1'] = "Toplam Miktar"
        ws['AT1'] = "Miktar Çoklama"
        ws['AU1'] = "Fark"
        
        for row in range(2, ws.max_row, 2):
            ws.merge_cells(f"AQ{row}:AQ{row+1}")
            ws[f"AQ{row}"] = f"=MAX(IF(SUM(S{row}:S{row+1})>0,IF(AH{row}=\"Muadil\",ROUNDUP(IFERROR(SUM(L{row}:L{row+1})/MAX(U{row}:U{row+1})*IF(AC{row+1}>0,AC{row+1},AK{row+1}),0),0)+S{row+1}+SUM(AB{row}:AB{row+1})-SUM(P{row}:P{row+1}),ROUNDUP((IFERROR(L{row}/U{row},0)+IFERROR(L{row+1}/U{row+1},0))*IF(AC{row+1}>0,AC{row+1},AK{row+1}),0)+S{row+1}+SUM(AB{row}:AB{row+1})-P{row+1}),0),0)"
            ws[f"AR{row}"] = f"=IF(AQ{row+1}<>\"\",AQ{row+1},AR{row})"
            ws.merge_cells(f"AS{row}:AS{row+1}")
            ws[f"AS{row}"] = f"=AP{row}+AP{row+1}"
            ws[f"AT{row}"] = f"=IF(AS{row+1}<>\"\",AS{row+1},AT{row})"
            ws[f"AU{row}"] = f"=AT{row}-AR{row}"
        
        wb.save(output_file)
        
        st.success("Dosya başarıyla işlendi!")
        
        # İşlenmiş dosyayı indirme bağlantısı oluşturma
        with open(output_file, "rb") as file:
            st.download_button("İşlenmiş dosyayı indir", file, file_name=output_file)
    except Exception as e:
        st.error(f"İşleme sırasında bir hata oluştu: {str(e)}")
